import streamlit as st
import pandas as pd
import openpyxl
import io, re
from openpyxl.styles import PatternFill

st.set_page_config(page_title="더존 전표 변환기", layout="wide")
st.title("📊 더존 위하고 전표 변환기")

# ---------------------------
# 기본 유틸
# ---------------------------
def to_int(v):
    try:
        if v is None:
            return 0
        if isinstance(v, str):
            v = re.sub(r"[^0-9.-]", "", v)
        return int(float(v))
    except:
        return 0

def clean(v):
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except:
        pass
    return str(v).strip()

def parse_date(v):
    try:
        d = pd.to_datetime(v)
        return d.month, d.day
    except:
        return None, None

# ---------------------------
# 🔥 안전 헤더 병합
# ---------------------------
def merge_headers(df, header_row, max_rows=3):
    parts = []

    max_row = min(len(df), header_row + max_rows)

    for i in range(header_row, max_row):
        row = df.iloc[i].fillna("").astype(str)
        parts.append(row)

    if not parts:
        return df.columns

    merged = []
    for col in range(len(parts[0])):
        vals = []
        for r in parts:
            if col < len(r):
                vals.append(str(r[col]).strip())

        name = " ".join([v for v in vals if v])
        merged.append(name if name else f"col_{col}")

    return merged

# ---------------------------
# 🔥 안전 DF 준비
# ---------------------------
def prepare_df(df):

    header_row = None
    for i in range(min(15, len(df))):
        row = df.iloc[i].astype(str)
        if row.str.contains("거래|일자|종목").any():
            header_row = i
            break

    if header_row is None:
        header_row = 0

    df.columns = merge_headers(df, header_row)

    # 데이터 시작 찾기
    start_row = header_row + 1
    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]
        if row.notna().sum() >= 3:
            start_row = i
            break

    df = df.iloc[start_row:].reset_index(drop=True)

    return df

# ---------------------------
# 🔥 패턴 탐지
# ---------------------------
def detect_column(df, check_fn, threshold=0.5):
    best_col, best_score = None, 0

    for col in df.columns:
        vals = df[col].dropna().head(30)
        if len(vals) == 0:
            continue

        score = sum(1 for v in vals if check_fn(v)) / len(vals)

        if score > best_score:
            best_score, best_col = score, col

    return best_col if best_score >= threshold else None

def is_date(v):
    try:
        pd.to_datetime(v)
        return True
    except:
        return False

def is_number(v):
    try:
        float(str(v).replace(",", ""))
        return True
    except:
        return False

def is_trade_type(v):
    v = str(v)
    return any(k in v for k in ["매수","매도","입금","출금","이체","입고","공모","배당","이자"])

# ---------------------------
# 스마트 컬럼 찾기
# ---------------------------
def smart_find(df, keywords, fn=None):
    for c in df.columns:
        if any(k in str(c) for k in keywords):
            return c

    if fn:
        col = detect_column(df, fn)
        if col:
            return col

    return df.columns[0]

# ---------------------------
# 🔥 증권사 감지
# ---------------------------
def detect_broker(df):
    sample = df.astype(str).fillna("").head(10).to_string()

    if "한국투자" in sample:
        return "HANTOO"
    if "키움" in sample:
        return "KIWOOM"
    if "삼성" in sample:
        return "SAMSUNG"
    if "미래에셋" in sample:
        return "MIRAE"

    return "UNKNOWN"

# ---------------------------
# 🔥 파서들
# ---------------------------
def parse_hantoo(df):

    df = prepare_df(df)

    def find(k):
        for c in df.columns:
            if k in str(c):
                return c
        return None

    c_date  = find("거래일")
    c_type  = find("구분")
    c_stock = find("종목")
    c_qty   = find("수량")
    c_price = find("단가")
    c_net   = find("금액")
    c_fee   = find("수수료")

    trades = []

    for _, r in df.iterrows():
        m,d = parse_date(r.get(c_date))
        if not m:
            continue

        trades.append({
            "month": m,
            "day": d,
            "type": clean(r.get(c_type)),
            "stock": clean(r.get(c_stock)),
            "qty": to_int(r.get(c_qty)),
            "price": to_int(r.get(c_price)),
            "net": to_int(r.get(c_net)),
            "fee": to_int(r.get(c_fee)),
            "tax": 0
        })

    return trades

def parse_generic(df):

    df = prepare_df(df)

    c_date  = smart_find(df, ["거래일","일자"], is_date)
    c_type  = smart_find(df, ["거래","구분"], is_trade_type)
    c_stock = smart_find(df, ["종목"])
    c_qty   = smart_find(df, ["수량"], is_number)
    c_price = smart_find(df, ["단가"], is_number)
    c_net   = smart_find(df, ["금액"], is_number)

    trades = []

    for _, r in df.iterrows():
        m,d = parse_date(r.get(c_date))
        if not m:
            continue

        trades.append({
            "month": m,
            "day": d,
            "type": clean(r.get(c_type)),
            "stock": clean(r.get(c_stock)),
            "qty": to_int(r.get(c_qty)),
            "price": to_int(r.get(c_price)),
            "net": to_int(r.get(c_net)),
            "fee": 0,
            "tax": 0
        })

    return trades

# ---------------------------
# 통합 파서
# ---------------------------
def parse_sheet(df):
    broker = detect_broker(df)

    if broker == "HANTOO":
        return parse_hantoo(df)

    return parse_generic(df)

# ---------------------------
# 거래처 매핑
# ---------------------------
def load_broker_map(file):
    if file is None:
        return {}

    df = pd.read_excel(file)

    return {
        str(name).strip(): (str(code).strip(), str(name).strip())
        for code, name in zip(df.iloc[:,0], df.iloc[:,1])
    }

def get_broker_info(stock, broker_map):
    return broker_map.get(stock, ("", "미등록거래처"))

# ---------------------------
# 거래유형
# ---------------------------
def normalize_trade_type(t):
    t = str(t)
    if "매도" in t: return "SELL"
    if "매수" in t: return "BUY"
    if "배당" in t: return "DIVIDEND"
    return None

# ---------------------------
# 전표 생성
# ---------------------------
def process_trades(trades, broker_map, broker_code, acc):

    rows = []

    for t in trades:
        ttype = normalize_trade_type(t["type"])
        if not ttype:
            continue

        m,d = t["month"], t["day"]
        qty, price = t["qty"], t["price"]
        net = t["net"]

        stock = t["stock"]
        cp_code, cp_name = get_broker_info(stock, broker_map)

        if ttype == "BUY":
            rows.append([m,d,"차변",acc["단기매매증권"],"단기매매증권",cp_code,cp_name,stock,qty*price,0])
            rows.append([m,d,"대변",acc["예치금"],"예치금",broker_code,"증권사",stock,0,net])

        elif ttype == "SELL":
            rows.append([m,d,"차변",acc["예치금"],"예치금",broker_code,"증권사",stock,net,0])
            rows.append([m,d,"대변",acc["단기매매증권"],"단기매매증권",cp_code,cp_name,stock,0,qty*price])

        elif ttype == "DIVIDEND":
            rows.append([m,d,"차변",acc["예치금"],"예치금",broker_code,"증권사","배당",net,0])
            rows.append([m,d,"대변",acc["배당금수익"],"배당금수익",cp_code,cp_name,"배당",0,net])

    return rows

# ---------------------------
# 검증
# ---------------------------
def validate_rows(rows):
    errors = []

    for i, r in enumerate(rows):
        dr, cr = r[8], r[9]

        if dr and cr:
            errors.append(f"{i+2}행 차변/대변 중복")

        if not dr and not cr:
            errors.append(f"{i+2}행 금액 없음")

    return errors

# ---------------------------
# 엑셀 생성
# ---------------------------
def create_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active

    header = ["월","일","구분","계정코드","계정명","거래처코드","거래처명","적요","차변","대변"]
    ws.append(header)

    fill = PatternFill(start_color="FFF59D", fill_type="solid")
    for col in range(1,len(header)+1):
        ws.cell(row=1,column=col).fill = fill

    for r in rows:
        ws.append(r)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ---------------------------
# UI
# ---------------------------
st.subheader("⚙️ 계정코드 설정")

acc = {
    "예치금": int(st.text_input("예치금","12500")),
    "단기매매증권": int(st.text_input("단기매매증권","10700")),
    "배당금수익": int(st.text_input("배당금수익","42100")),
}

broker_file = st.file_uploader("거래처 매핑")
broker_code = st.text_input("증권사 거래처코드")

uploaded = st.file_uploader("엑셀 업로드")

if uploaded and st.button("변환 실행"):

    broker_map = load_broker_map(broker_file)

    xl = pd.ExcelFile(uploaded)
    trades = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, header=None)
        trades += parse_sheet(df)

    rows = process_trades(trades, broker_map, broker_code, acc)

    errors = validate_rows(rows)

    if errors:
        st.error("❌ 오류 있음")
        for e in errors:
            st.write(e)
    else:
        st.success("완료")
        out = create_excel(rows)
        st.download_button("다운로드", data=out, file_name="result.xlsx")
